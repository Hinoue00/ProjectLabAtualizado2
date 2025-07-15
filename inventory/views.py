# inventory/views.py
from django.db.models import Count
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db import models
from django.db.models import Q
from accounts.views import is_technician
from .models import Material, MaterialCategory
from .forms import MaterialForm, MaterialCategoryForm, ImportMaterialsForm
from laboratories.models import Laboratory
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import csv
import io
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
import json
from .services import DoclingService
from django.conf import settings
from django.core.paginator import Paginator
from django.contrib import messages
from .automation_service import InventoryAutomationService
import os
import tempfile


docling_service = DoclingService() if getattr(settings, 'DOCLING_ENABLED', False) else None


@login_required
@user_passes_test(is_technician)
def material_list(request):
    """Lista de materiais com filtros e busca"""
    search_query = request.GET.get('search', '').strip()
    category_filter = request.GET.get('category', '')
    laboratory_filter = request.GET.get('laboratory', '')
    stock_status = request.GET.get('stock_status', '')
    
    # Query base otimizada
    materials = Material.objects.select_related('category', 'laboratory').all()
    
    # Aplicar filtros
    if search_query:
        materials = materials.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    if category_filter:
        materials = materials.filter(category_id=category_filter)
    
    if laboratory_filter:
        materials = materials.filter(laboratory_id=laboratory_filter)
    
    if stock_status == 'low':
        materials = materials.filter(is_low_stock=True)
    elif stock_status == 'normal':
        materials = materials.filter(is_low_stock=False)
    
    # Paginação eficiente
    paginator = Paginator(materials, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Categories e Labs com uma query só
    filter_data = {
        'categories': MaterialCategory.objects.all().order_by('name'),
        'laboratories': Laboratory.objects.filter(is_active=True).order_by('name')
    }
    
    context = {
        'materials': page_obj,
        'search_query': search_query,
        'category_filter': category_filter,
        'laboratory_filter': laboratory_filter,
        'stock_status': stock_status,
        **filter_data,
        'page_obj': page_obj,
        'paginator': paginator,
    }
    
    return render(request, 'material_list.html', context)

@login_required
@user_passes_test(is_technician)
def material_create(request):
    """Criar novo material"""
    if request.method == 'POST':
        form = MaterialForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Material adicionado com sucesso.')
            return redirect('material_list')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = MaterialForm()
    
    return render(request, 'material_form.html', {
        'form': form, 
        'title': 'Adicionar Material'
    })

@login_required
@user_passes_test(is_technician)
def material_update(request, pk):
    """Editar material existente"""
    material = get_object_or_404(Material, pk=pk)
    
    if request.method == 'POST':
        form = MaterialForm(request.POST, instance=material)
        if form.is_valid():
            form.save()
            messages.success(request, 'Material atualizado com sucesso.')
            return redirect('material_list')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = MaterialForm(instance=material)
    
    return render(request, 'material_form.html', {
        'form': form, 
        'title': 'Editar Material',
        'material': material
    })

@login_required
@user_passes_test(is_technician)
def material_delete(request, pk):
    """Deletar material"""
    material = get_object_or_404(Material, pk=pk)
    
    if request.method == 'POST':
        material_name = material.name
        material.delete()
        messages.success(request, f'Material "{material_name}" removido com sucesso.')
        return redirect('material_list')
    
    return render(request, 'material_confirm_delete.html', {
        'material': material,
        'title': 'Confirmar Exclusão'
    })

@login_required
def material_detail(request, pk):
    """Detalhes do material"""
    material = get_object_or_404(Material, pk=pk)
    
    context = {
        'material': material,
        'title': f'Detalhes - {material.name}'
    }
    
    return render(request, 'material_detail.html', context)

@login_required
@user_passes_test(is_technician)
def category_list(request):
    """Lista de categorias"""
    categories = MaterialCategory.objects.all().order_by('name')
    
    context = {
        'categories': categories,
        'title': 'Categorias de Materiais'
    }
    
    return render(request, 'category_list.html', context)

@login_required
@user_passes_test(is_technician)
def category_create(request):
    """Criar nova categoria"""
    if request.method == 'POST':
        form = MaterialCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoria criada com sucesso.')
            return redirect('category_list')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = MaterialCategoryForm()
    
    return render(request, 'category_form.html', {
        'form': form,
        'title': 'Nova Categoria'
    })

@login_required
@user_passes_test(is_technician)
def category_update(request, pk):
    """Editar categoria existente"""
    category = get_object_or_404(MaterialCategory, pk=pk)
    
    if request.method == 'POST':
        form = MaterialCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoria atualizada com sucesso.')
            return redirect('category_list')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = MaterialCategoryForm(instance=category)
    
    return render(request, 'category_form.html', {
        'form': form,
        'title': 'Editar Categoria',
        'category': category
    })

@login_required
@user_passes_test(is_technician)
def category_delete(request, pk):
    """Deletar categoria"""
    category = get_object_or_404(MaterialCategory, pk=pk)
    
    if request.method == 'POST':
        category_name = category.name
        category.delete()
        messages.success(request, f'Categoria "{category_name}" removida com sucesso.')
        return redirect('category_list')
    
    return render(request, 'category_confirm_delete.html', {
        'category': category,
        'title': 'Confirmar Exclusão'
    })

@login_required
@user_passes_test(is_technician)
def import_materials(request):
    """Importar materiais em lote - VERSÃO FINAL CORRIGIDA"""
    
    if request.method == 'POST':
        form = ImportMaterialsForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            
            # Opções de importação
            create_missing_categories = request.POST.get('create_missing_categories', False)
            create_missing_labs = request.POST.get('create_missing_labs', False)
            update_existing = request.POST.get('update_existing', False)
            skip_errors = request.POST.get('skip_errors', False)
            
            try:
                # Ler arquivo baseado na extensão
                if file.name.endswith('.csv'):
                    import pandas as pd
                    df = pd.read_csv(file)
                elif file.name.endswith(('.xlsx', '.xls')):
                    import pandas as pd
                    df = pd.read_excel(file)
                else:
                    messages.error(request, 'Formato de arquivo não suportado.')
                    return redirect('import_materials')
                
                # Debug: mostrar estrutura inicial
                print(f"DEBUG: Arquivo carregado com {len(df)} linhas")
                print(f"DEBUG: Colunas encontradas: {list(df.columns)}")
                if len(df) > 0:
                    print(f"DEBUG: Primeira linha: {df.iloc[0].tolist()}")
                if len(df) > 1:
                    print(f"DEBUG: Segunda linha: {df.iloc[1].tolist()}")
                
                # Verificar se a segunda linha contém cabeçalhos em português (linha de exemplo)
                if len(df) > 1:
                    second_row = df.iloc[1]
                    # Verificar se a segunda linha parece ser uma linha de exemplo/cabeçalho
                    if (pd.notna(second_row.iloc[0]) and 
                        isinstance(second_row.iloc[0], str) and 
                        ('Nome do Material' in str(second_row.iloc[0]) or 
                         'Descrição' in str(second_row.iloc[1]) if len(second_row) > 1 else False)):
                        print("DEBUG: Removendo linha de exemplo (linha 2)")
                        df = df.drop(df.index[1]).reset_index(drop=True)
                
                # Mapeamento de colunas (suporte a inglês e português)
                column_mapping = {
                    'nome': ['nome', 'name', 'material', 'NAME'],
                    'categoria': ['categoria', 'category', 'CATEGORY'],
                    'laboratorio': ['laboratorio', 'laboratório', 'laboratory', 'lab', 'LABORATORY'],
                    'quantidade': ['quantidade', 'quantity', 'qty', 'QUANTITY'],
                    'estoque_minimo': ['estoque_minimo', 'estoque_mínimo', 'minimum_stock', 'min_stock', 'MINIMUM_STOCK'],
                    'descricao': ['descricao', 'descrição', 'description', 'DESCRIPTION']
                }
                
                # Mapear colunas do arquivo
                df_columns = list(df.columns)
                mapped_columns = {}
                
                for standard_col, variations in column_mapping.items():
                    for variation in variations:
                        if variation in df_columns:
                            mapped_columns[standard_col] = variation
                            break
                
                print(f"DEBUG: Mapeamento de colunas: {mapped_columns}")
                
                # Verificar colunas obrigatórias
                required_columns = ['nome', 'categoria', 'laboratorio', 'quantidade', 'estoque_minimo']
                missing_cols = [col for col in required_columns if col not in mapped_columns]
                
                if missing_cols:
                    messages.error(request, f'Colunas obrigatórias não encontradas: {", ".join(missing_cols)}')
                    messages.error(request, f'Colunas disponíveis: {", ".join(df_columns)}')
                    return redirect('import_materials')
                
                # Renomear colunas para padronizar
                rename_dict = {v: k for k, v in mapped_columns.items()}
                df = df.rename(columns=rename_dict)
                
                print(f"DEBUG: Dados após renomeação: {df.columns.tolist()}")
                
                # Estatísticas
                imported_count = 0
                updated_count = 0
                skipped_count = 0
                errors = []
                
                # Processar cada linha
                for idx, row in df.iterrows():
                    try:
                        # Pular linhas vazias
                        if pd.isna(row.get('nome')) or str(row.get('nome')).strip() == '':
                            print(f"DEBUG: Pulando linha {idx + 1} - nome vazio")
                            skipped_count += 1
                            continue
                        
                        # Extrair dados
                        nome = str(row['nome']).strip()
                        categoria_nome = str(row['categoria']).strip()
                        laboratorio_nome = str(row['laboratorio']).strip()
                        
                        print(f"DEBUG: Processando linha {idx + 1} - {nome}")
                        
                        # Validar quantidade
                        try:
                            quantidade = int(float(row['quantidade']))
                            if quantidade < 0:
                                raise ValueError("Quantidade não pode ser negativa")
                        except (ValueError, TypeError):
                            error_msg = f"Linha {idx + 1}: Quantidade inválida - {row['quantidade']}"
                            errors.append(error_msg)
                            print(f"DEBUG: {error_msg}")
                            if not skip_errors:
                                continue
                            quantidade = 0
                        
                        # Validar estoque mínimo
                        try:
                            estoque_minimo = int(float(row['estoque_minimo']))
                            if estoque_minimo < 1:
                                estoque_minimo = 1
                        except (ValueError, TypeError):
                            error_msg = f"Linha {idx + 1}: Estoque mínimo inválido - {row['estoque_minimo']}"
                            errors.append(error_msg)
                            print(f"DEBUG: {error_msg}")
                            if not skip_errors:
                                continue
                            estoque_minimo = 1
                        
                        # Descrição opcional
                        descricao = ''
                        if 'descricao' in row and pd.notna(row['descricao']):
                            descricao = str(row['descricao']).strip()
                        
                        # Buscar ou criar categoria
                        categoria = None
                        try:
                            categoria = MaterialCategory.objects.get(name__iexact=categoria_nome)
                            print(f"DEBUG: Categoria encontrada: {categoria.name}")
                        except MaterialCategory.DoesNotExist:
                            if create_missing_categories:
                                categoria = MaterialCategory.objects.create(
                                    name=categoria_nome,
                                    material_type='consumable'
                                )
                                print(f"DEBUG: Categoria criada: {categoria.name}")
                            else:
                                error_msg = f"Linha {idx + 1}: Categoria '{categoria_nome}' não encontrada"
                                errors.append(error_msg)
                                print(f"DEBUG: {error_msg}")
                                if not skip_errors:
                                    continue
                        
                        # Buscar ou criar laboratório
                        laboratorio = None
                        try:
                            laboratorio = Laboratory.objects.get(name__iexact=laboratorio_nome)
                            print(f"DEBUG: Laboratório encontrado: {laboratorio.name}")
                        except Laboratory.DoesNotExist:
                            if create_missing_labs:
                                laboratorio = Laboratory.objects.create(
                                    name=laboratorio_nome,
                                    location=f"Localização {laboratorio_nome}",
                                    capacity=30,
                                    is_active=True
                                )
                                print(f"DEBUG: Laboratório criado: {laboratorio.name}")
                            else:
                                error_msg = f"Linha {idx + 1}: Laboratório '{laboratorio_nome}' não encontrado"
                                errors.append(error_msg)
                                print(f"DEBUG: {error_msg}")
                                if not skip_errors:
                                    continue
                        
                        # Pular se não temos categoria ou laboratório
                        if not categoria or not laboratorio:
                            skipped_count += 1
                            continue
                        
                        # Criar ou atualizar material
                        if update_existing:
                            material, created = Material.objects.update_or_create(
                                name=nome,
                                laboratory=laboratorio,
                                defaults={
                                    'category': categoria,
                                    'description': descricao,
                                    'quantity': quantidade,
                                    'minimum_stock': estoque_minimo,
                                }
                            )
                            if created:
                                imported_count += 1
                                print(f"DEBUG: Material criado: {material.name}")
                            else:
                                updated_count += 1
                                print(f"DEBUG: Material atualizado: {material.name}")
                        else:
                            # Verificar se já existe
                            if Material.objects.filter(name=nome, laboratory=laboratorio).exists():
                                error_msg = f"Linha {idx + 1}: Material '{nome}' já existe no laboratório '{laboratorio_nome}'"
                                errors.append(error_msg)
                                print(f"DEBUG: {error_msg}")
                                if not skip_errors:
                                    continue
                                skipped_count += 1
                                continue
                            
                            # Criar novo material
                            material = Material.objects.create(
                                name=nome,
                                category=categoria,
                                description=descricao,
                                quantity=quantidade,
                                minimum_stock=estoque_minimo,
                                laboratory=laboratorio,
                            )
                            imported_count += 1
                            print(f"DEBUG: Material criado: {material.name}")
                    
                    except Exception as e:
                        error_msg = f"Linha {idx + 1}: Erro ao processar - {str(e)}"
                        errors.append(error_msg)
                        print(f"DEBUG: {error_msg}")
                        if not skip_errors:
                            continue
                        skipped_count += 1
                
                # Mostrar resultados
                print(f"DEBUG: Resultado final - Importados: {imported_count}, Atualizados: {updated_count}, Pulados: {skipped_count}, Erros: {len(errors)}")
                
                success_messages = []
                if imported_count > 0:
                    success_messages.append(f"{imported_count} materiais importados")
                if updated_count > 0:
                    success_messages.append(f"{updated_count} materiais atualizados")
                
                if success_messages:
                    messages.success(request, f"Importação concluída: {', '.join(success_messages)}.")
                
                if skipped_count > 0:
                    messages.warning(request, f"{skipped_count} linhas foram ignoradas.")
                
                if errors:
                    messages.error(request, f"Encontrados {len(errors)} erros:")
                    for error in errors[:5]:  # Mostrar apenas os primeiros 5 erros
                        messages.error(request, error)
                    if len(errors) > 5:
                        messages.error(request, f"... e mais {len(errors) - 5} erros.")
                
                return redirect('material_list')
                
            except Exception as e:
                error_msg = f'Erro ao processar arquivo: {str(e)}'
                messages.error(request, error_msg)
                print(f"DEBUG: {error_msg}")
                import traceback
                traceback.print_exc()
                return redirect('import_materials')
    
    else:
        form = ImportMaterialsForm()
    
    # Contexto para o template
    existing_labs = Laboratory.objects.filter(is_active=True).values('name').distinct()
    existing_categories = MaterialCategory.objects.values('name').distinct()
    
    context = {
        'form': form,
        'title': 'Importar Materiais',
        'existing_labs': existing_labs,
        'existing_categories': existing_categories,
    }
    
    return render(request, 'import_materials.html', context)

@login_required
@user_passes_test(is_technician)
def export_materials(request):
    """Exportar materiais para Excel"""
    materials = Material.objects.select_related('category', 'laboratory').all()
    
    # Criar workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Materiais'
    
    # Cabeçalhos
    headers = [
        'Nome', 'Categoria', 'Tipo', 'Descrição', 
        'Quantidade', 'Estoque Mínimo', 'Laboratório', 'Status'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Adicionar dados
    for row_num, material in enumerate(materials, 2):
        worksheet.cell(row=row_num, column=1).value = material.name
        worksheet.cell(row=row_num, column=2).value = material.category.name
        worksheet.cell(row=row_num, column=3).value = material.category.get_material_type_display()
        worksheet.cell(row=row_num, column=4).value = material.description
        worksheet.cell(row=row_num, column=5).value = material.quantity
        worksheet.cell(row=row_num, column=6).value = material.minimum_stock
        worksheet.cell(row=row_num, column=7).value = material.laboratory.name
        
        stock_status = 'Estoque Baixo' if material.is_low_stock else 'OK'
        worksheet.cell(row=row_num, column=8).value = stock_status
    
    # Salvar em BytesIO
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    # Criar resposta
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=materiais_labconnect.xlsx'
    
    return response

@require_POST
def analyze_material_description(request):
    """API para analisar descrições de materiais em tempo real"""
    if not docling_service:
        return JsonResponse({'error': 'Serviço não disponível'}, status=400)
    
    try:
        data = json.loads(request.body)
        description = data.get('description', '')
        
        if not description or len(description) < 10:
            return JsonResponse({'error': 'Descrição muito curta'}, status=400)
        
        # Analisar o texto
        analysis = docling_service.analyze_text(description)
        
        # Obter a categoria sugerida
        suggested_category = docling_service.categorize_material(description)
        
        # Tentar encontrar a categoria no banco de dados
        category_id = None
        category_display = dict(Material.category.field.choices).get(suggested_category, 'Consumível')
        
        try:
            category = MaterialCategory.objects.filter(material_type=suggested_category).first()
            if category:
                category_id = category.id
                category_display = category.name
        except:
            pass
        
        return JsonResponse({
            'keywords': analysis['keywords'][:10],  # Limitar a 10 palavras-chave
            'suggested_category': suggested_category,
            'suggested_category_display': category_display,
            'suggested_category_id': category_id
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    
@login_required
@user_passes_test(is_technician)
def analyze_materials_batch(request):
    """Página para analisar materiais em lote"""
    if not docling_service:
        messages.warning(request, 'O serviço de análise docling não está disponível.')
        return redirect('material_list')
    
    # Contar materiais sem análise
    unanalyzed_count = Material.objects.filter(analyzed_data__isnull=True).count()
    
    if request.method == 'POST':
        # Análise em lote
        batch_size = int(request.POST.get('batch_size', 50))
        
        materials = Material.objects.filter(
            analyzed_data__isnull=True
        )[:batch_size]
        
        analyzed_count = 0
        for material in materials:
            if material.description:
                # Analisar
                material.analyzed_data = docling_service.analyze_text(material.description)
                material.suggested_category = docling_service.categorize_material(material.description)
                material.save()
                analyzed_count += 1
        
        if analyzed_count > 0:
            messages.success(request, f'{analyzed_count} materiais foram analisados com sucesso.')
        else:
            messages.info(request, 'Nenhum material foi analisado.')
            
        return redirect('analyze_materials_batch')
    
    return render(request, 'analyze_materials_batch.html', {
        'unanalyzed_count': unanalyzed_count
    })

@require_POST
@csrf_exempt
def suggest_material_details(request):
    """API para sugerir nomes e descrições para materiais"""
    if not docling_service:
        return JsonResponse({'error': 'Serviço não disponível'}, status=400)
    
    try:
        data = json.loads(request.body)
        partial_info = data.get('partial_info', '')
        category_id = data.get('category_id')
        
        if not partial_info or len(partial_info) < 3:
            return JsonResponse({'error': 'Informação insuficiente'}, status=400)
        
        # Obter a categoria se fornecida
        category_type = None
        if category_id:
            try:
                category = MaterialCategory.objects.get(id=category_id)
                category_type = category.material_type
            except:
                pass
        
        # Usar docling para gerar sugestões
        suggestions = docling_service.generate_material_suggestions(partial_info, category_type)
        
        return JsonResponse({
            'suggestions': suggestions
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    
@login_required
@user_passes_test(is_technician)
def material_trends(request):
    """Visualização de tendências e insights de materiais baseados na análise docling"""
    # Inicialize as variáveis antes do bloco condicional
    materials_with_analysis = list(Material.objects.exclude(analyzed_data__isnull=True).values('id', 'name')[:100])
    similar_materials = []

    if not docling_service:
        messages.warning(request, 'O serviço de análise docling não está disponível.')
        return redirect('material_list')

    # Contar materiais analisados
    analyzed_count = Material.objects.exclude(analyzed_data__isnull=True).count()
    total_count = Material.objects.count()
    analysis_percentage = (analyzed_count / total_count * 100) if total_count > 0 else 0

    # Agrupar por categoria sugerida
    category_distribution = Material.objects.exclude(
        suggested_category__isnull=True
    ).values('suggested_category').annotate(
        count=Count('id')
    ).order_by('-count')

    # Extrair palavras-chave comuns
    common_keywords = {}
    materials_with_keywords = Material.objects.exclude(
        analyzed_data__isnull=True
    )[:100]  # Limitar para não sobrecarregar

    for material in materials_with_keywords:
        if material.analyzed_data and 'keywords' in material.analyzed_data:
            for keyword in material.analyzed_data['keywords']:
                if keyword in common_keywords:
                    common_keywords[keyword] += 1
                else:
                    common_keywords[keyword] = 1

    # Ordenar as palavras-chave pela frequência
    sorted_keywords = sorted(common_keywords.items(), key=lambda x: x[1], reverse=True)[:20]

    # Encontrar materiais similares
    if 'material_id' in request.GET:
        try:
            material_id = int(request.GET['material_id'])
            source_material = Material.objects.get(id=material_id)

            if source_material.analyzed_data and 'keywords' in source_material.analyzed_data:
                source_keywords = set(source_material.analyzed_data['keywords'])

                # Encontrar materiais com palavras-chave em comum
                other_materials = Material.objects.exclude(id=material_id).exclude(
                    analyzed_data__isnull=True
                )[:50]  # Limitar para desempenho

                for material in other_materials:
                    if material.analyzed_data and 'keywords' in material.analyzed_data:
                        material_keywords = set(material.analyzed_data['keywords'])
                        common = source_keywords.intersection(material_keywords)

                        if len(common) > 0:
                            similarity = len(common) / max(len(source_keywords), len(material_keywords))

                            similar_materials.append({
                                'material': material,
                                'similarity': similarity,
                                'common_keywords': list(common)
                            })

                # Adicionar materiais para o formulário de materiais similares
                materials_with_analysis = Material.objects.exclude(
                    analyzed_data__isnull=True
                ).values('id', 'name')[:100]  # Limitar para desempenho

                # Ordenar por similaridade
                similar_materials = sorted(similar_materials, key=lambda x: x['similarity'], reverse=True)[:5]
        except Exception as e:
            # Adicione um log de erro ou tratamento de exceção mais específico
            print(f"Erro ao processar materiais similares: {e}")

    context = {
        'analyzed_count': analyzed_count,
        'total_count': total_count,
        'analysis_percentage': analysis_percentage,
        'category_distribution': category_distribution,
        'common_keywords': sorted_keywords,
        'similar_materials': similar_materials,
        'materials_with_analysis': materials_with_analysis,
    }

    return render(request, 'material_trends.html', context)

def _generate_csv_template(request):
    """
    Gera template CSV como fallback quando openpyxl não está disponível
    """
    import csv
    from datetime import datetime
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"LabConnect_Template_Materiais_{timestamp}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Adicionar BOM para Excel reconhecer UTF-8
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # Cabeçalho
    writer.writerow(['name', 'description', 'quantity', 'minimum_stock', 'category', 'laboratory'])
    
    # Exemplos
    examples = [
        ['Papel A4', 'Papel branco para impressão, tamanho A4', '500', '100', 'Material de Escritório', 'Laboratório Geral'],
        ['Microscópio Binocular', 'Microscópio para visualização de amostras', '3', '1', 'Equipamentos', 'Laboratório de Biologia'],
        ['Reagente Químico', 'Reagente para análises químicas', '25', '5', 'Reagentes', 'Laboratório de Química'],
        ['Computador Desktop', 'Computador para atividades administrativas', '10', '2', 'Informática', 'Laboratório de Informática'],
        ['Luvas Descartáveis', 'Luvas de proteção individual', '50', '10', 'Proteção', 'Laboratório Geral']
    ]
    
    for example in examples:
        writer.writerow(example)
    
    return response

@login_required
@user_passes_test(is_technician)
def download_template_excel(request):
    """
    Gera e faz download de template Excel para importação de materiais
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Template_Materiais'
        
        # Definir estilo de borda
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Cabeçalhos principais
        headers = [
            'name', 'description', 'quantity', 'minimum_stock', 
            'category', 'laboratory'
        ]
        
        # Descrições dos cabeçalhos
        header_descriptions = [
            'Nome do Material',
            'Descrição Detalhada', 
            'Quantidade Atual',
            'Estoque Mínimo',
            'Categoria do Material',
            'Laboratório de Destino'
        ]
        
        # Aplicar cabeçalhos com formatação avançada
        for col, (header, description) in enumerate(zip(headers, header_descriptions), 1):
            # Cabeçalho principal (linha 1)
            cell = ws.cell(row=1, column=col, value=header.upper())
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
            # Descrição (linha 2)
            desc_cell = ws.cell(row=2, column=col, value=description)
            desc_cell.font = Font(italic=True, size=10, color="444444")
            desc_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            desc_cell.alignment = Alignment(horizontal="center", vertical="center")
            desc_cell.border = thin_border
        
        # Dados de exemplo com diferentes tipos de materiais
        examples = [
            # [nome, descrição, quantidade, estoque_min, categoria, laboratório]
            [
                'Papel A4 Sulfite',
                'Papel branco para impressão e documentos, tamanho A4, 75g/m²',
                500, 100,
                'Material de Escritório',
                'Laboratório Geral'
            ],
            [
                'Microscópio Binocular Zeiss',
                'Microscópio binocular para visualização de amostras microscópicas, objetivas 10x, 40x, 100x',
                3, 1,
                'Equipamentos de Laboratório',
                'Laboratório de Biologia'
            ],
            [
                'Ácido Clorídrico 37%',
                'Reagente químico para análises qualitativas e quantitativas, pureza analítica',
                25, 5,
                'Reagentes Químicos',
                'Laboratório de Química'
            ],
            [
                'Computador Desktop Dell',
                'Computador para atividades administrativas e ensino, Intel i5, 8GB RAM, SSD 256GB',
                10, 2,
                'Equipamentos de Informática',
                'Laboratório de Informática'
            ],
            [
                'Luvas Nitrilo Descartáveis',
                'Luvas de proteção individual, sem talco, tamanho M, caixa com 100 unidades',
                50, 10,
                'Equipamentos de Proteção',
                'Laboratório Geral'
            ],
            [
                'Pipeta Automática 100-1000μL',
                'Pipeta de volume variável para medições precisas, certificada e calibrada',
                8, 2,
                'Instrumentos de Medição',
                'Laboratório de Análises'
            ]
        ]
        
        # Adicionar exemplos com formatação
        for row_idx, example in enumerate(examples, 3):  # Começar na linha 3
            for col_idx, value in enumerate(example, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Formatação específica por tipo de dado
                if col_idx in [3, 4]:  # Colunas de quantidade
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(bold=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Ajustar largura das colunas automaticamente
        column_widths = [25, 50, 12, 15, 30, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Ajustar altura das linhas
        ws.row_dimensions[1].height = 25  # Cabeçalho
        ws.row_dimensions[2].height = 20  # Descrição
        
        # Adicionar seção de instruções
        instructions_start_row = len(examples) + 5
        
        # Título das instruções
        title_cell = ws.cell(row=instructions_start_row, column=1, value="📋 INSTRUÇÕES DE USO")
        title_cell.font = Font(bold=True, size=14, color="2F5597")
        title_cell.fill = PatternFill(start_color="E8F1FF", end_color="E8F1FF", fill_type="solid")
        
        # Mesclar células para o título
        ws.merge_cells(f'A{instructions_start_row}:F{instructions_start_row}')
        
        # Instruções detalhadas
        instructions = [
            "",  # Linha em branco
            "✅ COLUNAS OBRIGATÓRIAS:",
            "   • name: Nome do material (único e descritivo)",
            "   • quantity: Quantidade atual em estoque (número)",
            "   • minimum_stock: Estoque mínimo para alerta (número)",
            "",
            "📝 COLUNAS OPCIONAIS:",
            "   • description: Descrição detalhada (recomendado para IA)",
            "   • category: Categoria do material (será sugerida automaticamente)",
            "   • laboratory: Laboratório de destino (será atribuído automaticamente)",
            "",
            "🤖 RECURSOS INTELIGENTES:",
            "   • O sistema LabConnect usa IA para categorizar materiais automaticamente",
            "   • Descrições detalhadas melhoram a precisão da categorização",
            "   • Materiais similares são detectados para evitar duplicatas",
            "",
            "📊 DICAS DE PREENCHIMENTO:",
            "   • Use nomes claros e específicos (ex: 'Papel A4' em vez de 'Papel')",
            "   • Inclua marca/modelo quando relevante",
            "   • Especifique unidades de medida na descrição",
            "   • Defina estoques mínimos realistas para cada material",
            "",
            "🚀 APÓS PREENCHER:",
            "   1. Salve o arquivo Excel",
            "   2. Acesse LabConnect > Inventário > Importar Materiais",
            "   3. Faça upload do arquivo preenchido",
            "   4. Aguarde o processamento automático"
        ]
        
        # Adicionar instruções
        for i, instruction in enumerate(instructions, 1):
            instruction_cell = ws.cell(row=instructions_start_row + i, column=1, value=instruction)
            
            if instruction.startswith("✅") or instruction.startswith("📝") or instruction.startswith("🤖") or instruction.startswith("📊") or instruction.startswith("🚀"):
                instruction_cell.font = Font(bold=True, size=11, color="2F5597")
            elif instruction.startswith("   •") or instruction.startswith("   "):
                instruction_cell.font = Font(size=10, color="555555")
            else:
                instruction_cell.font = Font(size=10)
            
            # Mesclar células para instruções
            if instruction:
                ws.merge_cells(f'A{instructions_start_row + i}:F{instructions_start_row + i}')
        
        # Adicionar informações do sistema no rodapé
        footer_row = instructions_start_row + len(instructions) + 2
        footer_cell = ws.cell(row=footer_row, column=1, value="📱 LabConnect - Sistema de Gestão Laboratorial | Template gerado automaticamente")
        footer_cell.font = Font(size=9, italic=True, color="888888")
        footer_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(f'A{footer_row}:F{footer_row}')
        
        # Salvar em BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Criar resposta HTTP com nome de arquivo personalizado
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"LabConnect_Template_Materiais_{timestamp}.xlsx"
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Log da ação
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Template Excel gerado por {request.user.username}")
        
        return response
        
    except ImportError:
        # Fallback para CSV se openpyxl não estiver disponível
        messages.warning(request, 'Excel não disponível. Gerando template CSV.')
        return _generate_csv_template(request)
        
    except Exception as e:
        messages.error(request, f'Erro ao gerar template: {str(e)}')
        return redirect('import_materials')

@login_required
@user_passes_test(is_technician)
def automated_import_materials(request):
    """
    View simplificada para importação automatizada de materiais
    Por enquanto, redireciona para a importação normal até implementar a automação completa
    """
    messages.info(request, 'Funcionalidade de automação em desenvolvimento. Usando importação padrão.')
    return redirect('import_materials')

@login_required
@user_passes_test(is_technician)
def import_results(request):
    """
    Exibe resultados da importação automatizada
    """
    # Por enquanto, apenas uma página simples
    context = {
        'message': 'Funcionalidade em desenvolvimento'
    }
    return render(request, 'inventory/import_results_placeholder.html', context)

@csrf_exempt
@require_POST
def validate_file_ajax(request):
    """
    Valida arquivo via AJAX antes do upload
    """
    # Implementação básica por enquanto
    if 'file' not in request.FILES:
        return JsonResponse({'valid': False, 'error': 'Nenhum arquivo enviado'})
    
    uploaded_file = request.FILES['file']
    
    # Validação básica
    if not uploaded_file.name.endswith(('.xlsx', '.xls', '.csv')):
        return JsonResponse({
            'valid': False, 
            'error': 'Formato de arquivo inválido'
        })
    
    return JsonResponse({
        'valid': True,
        'message': 'Arquivo válido',
        'stats': {
            'total_rows': 0,  # Será implementado
            'columns_found': [],
            'required_columns_missing': [],
            'optional_columns_missing': []
        }
    })

@csrf_exempt
@require_POST
def preview_materials_ajax(request):
    """
    Prévia dos materiais que serão importados
    """
    # Implementação básica
    return JsonResponse({
        'success': True,
        'preview_data': [],
        'total_rows': 0,
        'preview_rows': 0
    })

@login_required
@user_passes_test(is_technician)
def bulk_categorization(request):
    """
    Categorização em lote de materiais existentes
    """
    if request.method == 'POST':
        messages.info(request, 'Funcionalidade de categorização em lote em desenvolvimento.')
        return redirect('material_list')
    
    context = {
        'materials_count': 0,  # Será implementado depois
        'automation_enabled': False
    }
    
    return render(request, 'inventory/bulk_categorization_placeholder.html', context)

@csrf_exempt
@require_POST
def auto_suggest_material(request):
    """
    Sugestão automática de material baseado em entrada parcial
    """
    try:
        data = json.loads(request.body)
        partial_input = data.get('input', '')
        
        # Implementação básica - retorna sugestões vazias por enquanto
        return JsonResponse({
            'success': True,
            'suggestions': []
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })